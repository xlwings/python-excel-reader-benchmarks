import itertools
import numbers
import os
import platform
import sys
import timeit
from pathlib import Path

import openpyxl
import pandas as pd
import psutil
import pyxlsb
import xlrd
import xlwings as xw
from psutil._common import bytes2human
from xlwings.utils import a1_to_tuples


# xlwings
def xlwings_get_sheet_values():
    with xw.Book(TEST_FILE, mode="r") as book:
        return book.sheets[SHEET].cells.value


def xlwings_get_range_values():
    with xw.Book(TEST_FILE, mode="r") as book:
        return book.sheets[SHEET][ADDRESS].value


def xlwings_get_sheet_df():
    with xw.Book(TEST_FILE, mode="r") as book:
        return book.sheets[SHEET].cells.options("df", index=False).value


def xlwings_get_range_df():
    with xw.Book(TEST_FILE, mode="r") as book:
        return book.sheets[SHEET][ADDRESS].options("df", index=False, header=False).value


# pandas
def pandas_get_sheet_df():
    return pd.read_excel(TEST_FILE, sheet_name=SHEET)


def pandas_get_range_df():
    return pd.read_excel(
        TEST_FILE,
        sheet_name=SHEET,
        usecols=list(range(COL1 - 1, COL2)),
        skiprows=ROW1 - 1,
        nrows=ROW2 - ROW1 + 1,
        header=None,
    )


# openpyxl
def openpyxl_get_sheet_values():
    book = openpyxl.load_workbook(
        TEST_FILE, read_only=True, keep_links=False, data_only=True
    )
    sheet = book.worksheets[SHEET]
    values = []
    for row in sheet.iter_rows(values_only=True):
        values.append(row)
    return values


def openpyxl_get_range_values():
    book = openpyxl.load_workbook(
        TEST_FILE, read_only=True, keep_links=False, data_only=True
    )
    sheet = book.worksheets[SHEET]
    values = []
    for row in sheet.iter_rows(
        min_row=ROW1, max_row=ROW2, min_col=COL1, max_col=COL2, values_only=True
    ):
        values.append(row)
    return values


# pyxlsb
def pyxlsb_get_sheet_values():
    values = []
    with pyxlsb.open_workbook(TEST_FILE) as book:
        with book.get_sheet(SHEET + 1) as sheet:
            for row in sheet.rows():
                values.append([cell.v for cell in row])
    return values


def pyxlsb_get_range_values():
    values = []
    with pyxlsb.open_workbook(TEST_FILE) as book:
        with book.get_sheet(SHEET + 1) as sheet:
            for row in itertools.islice(sheet.rows(), ROW1 - 1, ROW2):
                values.append([cell.v for cell in row][COL1 - 1 : COL2])
    return values


# xlrd
def xlrd_get_sheet_values():
    with xlrd.open_workbook(TEST_FILE, on_demand=True) as book:
        sheet = book.sheet_by_index(SHEET)
        return [sheet.row_values(row) for row in range(sheet.nrows)]


def xlrd_get_range_values():
    with xlrd.open_workbook(TEST_FILE, on_demand=True) as book:
        sheet = book.sheet_by_index(SHEET)
        return [
            sheet.row_values(row, start_colx=COL1 - 1, end_colx=COL2)
            for row in range(ROW1 - 1, ROW2)
        ]


def compare(func_one, func_two):
    one = func_one()
    two = func_two()

    if func_one.__name__.split("_")[1:] != func_two.__name__.split("_")[1:]:
        raise Exception("You're comparing different functions!")

    if isinstance(one, list) and not isinstance(one[0], list):
        raise Exception("Only single cells or 2d ranges are supported for address!")

    # DataFrame
    if isinstance(one, pd.DataFrame):
        # Align data
        for col in two.columns:
            if two[col].dtype == int:
                two[col] = two[col].astype(float)
        # Compare
        if one.equals(two):
            return
        else:
            print(two[~two.isin(one)].dropna())
            raise Exception("Values differ, see diff above.")

    # Single cells
    if not isinstance(one, list):
        two = two[0][0]
        two = None if two == "" else two
    else:
        # Lists of lists/tuples
        # Align Data
        two = [list(row) for row in two]
        two = [[None if cell == "" else cell for cell in row] for row in two]
        two = [
            [float(cell) if isinstance(cell, numbers.Number) else cell for cell in row]
            for row in two
        ]
    # Compare
    if one == two:
        return
    else:
        if not isinstance(one, list):
            raise Exception(f"Value differs: {one} vs. {two}")
        for ix, row in enumerate(one):
            if row != two[ix]:
                print(f"Excel Row: {ix + 1}")
                print(row)
                print(two[ix])
        raise Exception("Values differ, see diff above.")


def main(func_one, func_two, repeat, loops, description=None):
    module_one_name = func_one.__name__.split("_")[0]
    module_two_name = func_two.__name__.split("_")[0]
    time_one = min(timeit.repeat(func_one, repeat=repeat, number=loops)) / loops
    time_two = min(timeit.repeat(func_two, repeat=repeat, number=loops)) / loops

    compare(func_one, func_two)
    speedup = time_two / time_one

    print("=" * 80)
    print(
        f"[{Path(TEST_FILE).suffix[1:]}|{module_two_name}] {description}"
        if description
        else f"[{Path(TEST_FILE).suffix[1:]}|{module_two_name}] (no description)"
    )
    print("=" * 80)
    print(f"{func_one.__name__} vs. {func_two.__name__}")
    print(
        f"File: {TEST_FILE}, Sheet: {SHEET}, Address: {ADDRESS if ADDRESS else 'full sheet'}, "
        f"Repeat: {repeat}, Loops: {loops}"
    )
    print(" " * 80)
    print(f"{module_one_name}: {time_one:.3f}s")
    print(f"{module_two_name}: {time_two:.3f}s")
    print(f"Speedup {module_one_name} vs. {module_two_name}: {speedup:.1f}x")
    print("=" * 80)
    print()


if __name__ == "__main__":
    print(f"Python: {sys.version.split()[0]}")
    print(f"xlwings: {xw.__version__}")
    print(f"OpenPyXL: {openpyxl.__version__}")
    print(f"pyxlsb: {pyxlsb.__version__}")
    print(f"xlrd: {xlrd.__version__}")
    print(f"pandas: {pd.__version__}")
    print()
    print(f"Available Memory: {bytes2human(psutil.virtual_memory().available)}")
    print(f"CPUs: {os.cpu_count()}")
    print(f"Platform: {sys.platform}")
    print(f"Processor: {platform.processor()}")
    print()

    test_cases = (
        {
            "description": "sheet (10,500 rows)",
            "file": "xl/AAPL.xlsx",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_df,
            "two": pandas_get_sheet_df,
        },
        {
            "description": "top 10 rows from 10.k rows",
            "file": "xl/AAPL.xlsx",
            "sheet": 0,
            "address": "A1:G10",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_df,
            "two": pandas_get_range_df,
        },
        {
            "description": "bottom 10 rows from 10.k rows",
            "file": "xl/AAPL.xlsx",
            "sheet": 0,
            "address": "A10544:G10553",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_df,
            "two": pandas_get_range_df,
        },
        {
            "description": "small file, small df",
            "file": "xl/small.xlsx",
            "sheet": 0,
            "address": "A1:C3",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_df,
            "two": pandas_get_range_df,
        },
        {
            "description": "Read sheet (10,500 rows)",
            "file": "xl/AAPL.xlsx",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_values,
            "two": openpyxl_get_sheet_values,
        },
        {
            "description": "Read cell at top of 10,500 rows",
            "file": "xl/AAPL.xlsx",
            "sheet": 0,
            "address": "A1",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_values,
            "two": openpyxl_get_range_values,
        },
        {
            "description": "Read cell in row 10,000 of 10,500 rows",
            "file": "xl/AAPL.xlsx",
            "sheet": 0,
            "address": "D10000",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_values,
            "two": openpyxl_get_range_values,
        },
        {
            "description": "Read sheet in small file",
            "file": "xl/small.xlsx",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_values,
            "two": openpyxl_get_sheet_values,
        },
        {
            "description": "sheet (10,500 rows)",
            "file": "xl/AAPL.xlsb",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_df,
            "two": pandas_get_sheet_df,
        },
        {
            "description": "top 10 rows from 10.k rows",
            "file": "xl/AAPL.xlsb",
            "sheet": 0,
            "address": "A1:G10",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_df,
            "two": pandas_get_range_df,
        },
        {
            "description": "bottom 10 rows from 10.k rows",
            "file": "xl/AAPL.xlsb",
            "sheet": 0,
            "address": "A10544:G10553",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_df,
            "two": pandas_get_range_df,
        },
        {
            "description": "small file, small df",
            "file": "xl/small.xlsb",
            "sheet": 0,
            "address": "A1:C3",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_df,
            "two": pandas_get_range_df,
        },
        {
            "description": "Read sheet (10,500 rows)",
            "file": "xl/AAPL.xlsb",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_values,
            "two": pyxlsb_get_sheet_values,
        },
        {
            "description": "Read cell at top of 10,500 rows",
            "file": "xl/AAPL.xlsb",
            "sheet": 0,
            "address": "A1",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_values,
            "two": pyxlsb_get_range_values,
        },
        {
            "description": "Read cell in row 10,000 of 10,500 rows",
            "file": "xl/AAPL.xlsb",
            "sheet": 0,
            "address": "D10000",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_values,
            "two": pyxlsb_get_range_values,
        },
        {
            "description": "Read sheet in small file",
            "file": "xl/small.xlsb",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_values,
            "two": pyxlsb_get_sheet_values,
        },
        {
            "description": "Read sheet (10,500 rows)",
            "file": "xl/AAPL.xls",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_values,
            "two": xlrd_get_sheet_values,
        },
        {
            "description": "Read cell at top of 10,500 rows",
            "file": "xl/AAPL.xls",
            "sheet": 0,
            "address": "A1",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_values,
            "two": xlrd_get_range_values,
        },
        {
            "description": "Read cell in row 10,000 of 10,500 rows",
            "file": "xl/AAPL.xls",
            "sheet": 0,
            "address": "D10000",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_range_values,
            "two": xlrd_get_range_values,
        },
        {
            "description": "Read sheet in small file",
            "file": "xl/small.xls",
            "sheet": 0,
            "address": "",
            "repeat": 5,
            "loops": 10,
            "one": xlwings_get_sheet_values,
            "two": xlrd_get_sheet_values,
        },
    )

    for test in test_cases:
        TEST_FILE = test["file"]
        SHEET = test["sheet"]
        ADDRESS = test.get("address")
        if ADDRESS:
            cell1, cell2 = a1_to_tuples(ADDRESS)  # 1-based
            if not cell2:
                cell2 = cell1
            ROW1, COL1 = cell1[0], cell1[1]
            ROW2, COL2 = cell2[0], cell2[1]

        main(
            test["one"],
            test["two"],
            repeat=test["repeat"],
            loops=test["loops"],
            description=test.get("description"),
        )
